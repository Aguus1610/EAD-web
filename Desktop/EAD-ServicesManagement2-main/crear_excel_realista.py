"""
Script para crear archivos Excel más realistas con diferentes formatos
para probar el parser mejorado
"""
import pandas as pd
from datetime import datetime, timedelta
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def create_realistic_excel():
    """Crear archivo Excel con formatos más realistas y variados"""
    
    # Crear workbook
    wb = openpyxl.Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # Cliente 1: Formato con encabezados en fila 3
    ws1 = wb.create_sheet("TRANSPORTES RODRIGUEZ")
    
    # Título y información del cliente
    ws1['A1'] = "TRANSPORTES RODRIGUEZ S.A."
    ws1['A1'].font = Font(size=14, bold=True)
    ws1['A2'] = "Historial de Mantenimientos - 2024"
    
    # Encabezados en fila 4
    headers1 = ['VEHICULO', 'FECHA SERVICIO', 'REPUESTOS UTILIZADOS', 'MANO DE OBRA REALIZADA']
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Datos del cliente 1
    data1 = [
        ['SCANIA R450 Patente ABC123', '15/01/2024', 'Filtro aceite motor, Aceite 15W40 20L - $28000', 'Cambio aceite y filtros, Revisión general - $18000'],
        ['SCANIA R450 Patente ABC123', '20/03/2024', 'Pastillas freno delanteras Ferodo - $52000', 'Cambio pastillas, Rectificado discos - $25000'],
        ['VOLVO FH16 Patente DEF456', '10/02/2024', 'Filtro aire, Aceite hidráulico Shell - $22000', 'Service 10000km completo - $30000'],
        ['VOLVO FH16 Patente DEF456', '15/04/2024', 'Kit distribución Gates, Tensor - $85000', 'Cambio distribución completa - $65000'],
        ['MERCEDES ACTROS Patente GHI789', '05/03/2024', 'Amortiguadores traseros x2 - $95000', 'Cambio amortiguadores y alineación - $35000']
    ]
    
    for row, data_row in enumerate(data1, 5):
        for col, value in enumerate(data_row, 1):
            ws1.cell(row=row, column=col, value=value)
    
    # Cliente 2: Formato con columnas diferentes y espacios
    ws2 = wb.create_sheet("CONSTRUCCIONES MARTINEZ")
    
    # Información del cliente con espacios
    ws2['B2'] = "CONSTRUCCIONES MARTINEZ"
    ws2['B2'].font = Font(size=12, bold=True)
    
    # Encabezados en fila 5 con nombres diferentes
    headers2 = ['MAQUINA', 'DIA', 'MATERIALES Y REPUESTOS', 'TRABAJOS REALIZADOS']
    for col, header in enumerate(headers2, 2):  # Empezar en columna B
        cell = ws2.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # Datos del cliente 2
    data2 = [
        ['CATERPILLAR 320D Serie CAT001', '05-01-2024', 'Filtros hidráulicos x3 Caterpillar, Aceite hidráulico 68 - $75000', 'Service 500 horas, Revisión sistema hidráulico - $45000'],
        ['CATERPILLAR 320D Serie CAT001', '20-02-2024', 'Cadenas oruga nuevas, Pines y bujes - $180000', 'Cambio cadenas completo, Ajuste tensión - $120000'],
        ['JOHN DEERE 6120 Año 2019', '12-01-2024', 'Filtro combustible, Aceite motor Mobil 1 - $35000', 'Service preventivo 200hs - $28000'],
        ['KOMATSU PC200 Serie KOM789', '25-03-2024', 'Cilindro hidráulico brazo - $125000', 'Reparación cilindro, Pruebas sistema - $85000']
    ]
    
    for row, data_row in enumerate(data2, 6):
        for col, value in enumerate(data_row, 2):  # Empezar en columna B
            ws2.cell(row=row, column=col, value=value)
    
    # Cliente 3: Formato más desordenado con celdas mezcladas
    ws3 = wb.create_sheet("AGRO SAN MARTIN")
    
    # Título disperso
    ws3['A1'] = "ESTABLECIMIENTO AGROPECUARIO"
    ws3['A2'] = "SAN MARTIN"
    ws3['A3'] = "Registro de Mantenimientos"
    
    # Encabezados en diferentes posiciones
    ws3['A6'] = 'EQUIPO AGRICOLA'
    ws3['B6'] = 'CUANDO'
    ws3['C6'] = 'REPUESTOS'
    ws3['D6'] = 'SERVICIO'
    
    # Hacer encabezados visibles
    for cell in ['A6', 'B6', 'C6', 'D6']:
        ws3[cell].font = Font(bold=True)
        ws3[cell].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    # Datos del cliente 3 con formatos de fecha variados
    data3 = [
        ['MASSEY FERGUSON 5650', '8/1/2024', 'Filtro aire K&N, Aceite motor Castrol - $32000', 'Service 200 horas completo - $25000'],
        ['MASSEY FERGUSON 5650', '15/02/24', 'Kit embrague completo Sachs - $110000', 'Cambio embrague, Ajuste pedal - $75000'],
        ['NEW HOLLAND T7060', '22-ENE-2024', 'Filtros varios, Aceites Shell - $38000', 'Service completo, Engrase - $30000'],
        ['DEUTZ FAHR 6160', '30/1/2024', 'Bomba hidráulica Bosch reparada - $95000', 'Instalación bomba, Calibración - $55000'],
        ['CASE IH PUMA 180', '2024-02-14', 'Neumáticos Michelin 710/70R42 x2 - $150000', 'Cambio neumáticos traseros - $35000']
    ]
    
    for row, data_row in enumerate(data3, 7):
        for col, value in enumerate(data_row, 1):
            ws3.cell(row=row, column=col, value=value)
    
    # Cliente 4: Formato con tabla más formal pero nombres de columnas diferentes
    ws4 = wb.create_sheet("MINERA NORTE")
    
    ws4['A1'] = "MINERA DEL NORTE S.A."
    ws4['A1'].font = Font(size=14, bold=True)
    ws4['A2'] = "Control de Mantenimiento Equipos Pesados"
    
    # Encabezados más formales
    headers4 = ['UNIDAD', 'FECHA_TRABAJO', 'COMPONENTES_REPUESTOS', 'LABOR_TECNICA']
    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    # Datos del cliente 4
    data4 = [
        ['LIEBHERR R996 Excavadora', '2024/01/10', 'Filtros hidráulicos Liebherr x5, Aceite Mobil DTE 25 - $125000', 'Service 1000 horas, Análisis aceite - $85000'],
        ['CATERPILLAR 797F Dumper', '2024/01/25', 'Neumáticos Michelin 59/80R63 x1 - $280000', 'Cambio neumático dañado - $45000'],
        ['KOMATSU PC8000 Pala', '2024/02/08', 'Motor diesel Cummins QSK78 reparado - $450000', 'Instalación motor, Puesta en marcha - $180000'],
        ['HITACHI EX5600 Excavadora', '2024/02/20', 'Sistema de aire acondicionado - $85000', 'Reparación A/C cabina - $35000']
    ]
    
    for row, data_row in enumerate(data4, 5):
        for col, value in enumerate(data_row, 1):
            ws4.cell(row=row, column=col, value=value)
    
    # Guardar archivo
    wb.save('datos_realistas.xlsx')
    print("Archivo 'datos_realistas.xlsx' creado exitosamente!")
    print("Características del archivo:")
    print("- 4 hojas con diferentes formatos")
    print("- Encabezados en diferentes filas")
    print("- Nombres de columnas variados")
    print("- Formatos de fecha diversos")
    print("- Datos con información realista")

if __name__ == "__main__":
    create_realistic_excel()
