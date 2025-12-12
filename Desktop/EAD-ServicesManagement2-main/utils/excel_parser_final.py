"""
Parser Excel Final - Interpreta correctamente la estructura real del Excel
"""
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pd = None

import openpyxl
from datetime import date, datetime
import re
import logging
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass

logger = logging.getLogger(__name__)

@dataclass
class WorkEntry:
    fecha: date
    descripcion: str
    presupuesto: float = 0.0  # Cambiar a presupuesto para compatibilidad

@dataclass
class EquipmentData:
    propietario: str
    nombre: str
    trabajos: List[WorkEntry]

class ExcelParserFinal:
    def __init__(self):
        self.errors = []
        self.warnings = []
    
    def parse_excel_file(self, file_path: str) -> List[EquipmentData]:
        """Parsear archivo Excel con la estructura real identificada"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            all_equipment = []
            
            for sheet_name in workbook.sheetnames:
                if sheet_name.lower() in ['hoja5', 'sheet5']:  # Saltar hojas vacías
                    continue
                    
                logger.info(f"Procesando hoja: {sheet_name}")
                equipment_data = self._process_sheet_openpyxl(workbook, sheet_name)
                all_equipment.extend(equipment_data)
            
            workbook.close()
            return all_equipment
            
        except Exception as e:
            logger.error(f"Error procesando archivo Excel: {e}")
            self.errors.append(f"Error general: {e}")
            return []
    
    def _process_sheet_openpyxl(self, workbook, sheet_name: str) -> List[EquipmentData]:
        """Procesar una hoja usando openpyxl para mayor control"""
        try:
            worksheet = workbook[sheet_name]
            equipment_list = []
            
            # Encontrar la fila de headers (normalmente fila 3)
            header_row = self._find_header_row(worksheet)
            if not header_row:
                logger.warning(f"No se encontraron headers en '{sheet_name}'")
                return []
            
            logger.info(f"Headers encontrados en fila {header_row} de '{sheet_name}'")
            
            # Procesar datos a partir de la fila siguiente a los headers
            current_equipment = None
            current_date = None
            current_repuestos = []
            current_mano_obra = []
            
            max_row = worksheet.max_row
            
            for row_num in range(header_row + 1, max_row + 1):
                # Leer valores de la fila
                equipo_cell = worksheet.cell(row=row_num, column=1).value
                fecha_cell = worksheet.cell(row=row_num, column=2).value
                repuesto_cell = worksheet.cell(row=row_num, column=3).value
                mano_obra_cell = worksheet.cell(row=row_num, column=4).value
                
                # Limpiar valores
                equipo_val = self._clean_string(equipo_cell) if equipo_cell else None
                fecha_val = self._parse_date(fecha_cell) if fecha_cell else None
                repuesto_val = self._clean_string(repuesto_cell) if repuesto_cell else None
                mano_obra_val = self._clean_string(mano_obra_cell) if mano_obra_cell else None
                
                # Verificar si es una nueva entrada de equipo
                if equipo_val and equipo_val.upper() not in ['EQUIPO', 'EQUIPOS']:
                    # Guardar trabajo anterior si existe
                    if current_equipment and (current_repuestos or current_mano_obra):
                        trabajo = self._create_work_entry(current_date, current_repuestos, current_mano_obra)
                        if trabajo:
                            # Buscar si ya existe este equipo
                            existing_eq = None
                            for eq in equipment_list:
                                if eq.nombre == current_equipment:
                                    existing_eq = eq
                                    break
                            
                            if existing_eq:
                                existing_eq.trabajos.append(trabajo)
                            else:
                                equipment_list.append(EquipmentData(
                                    propietario=sheet_name,
                                    nombre=current_equipment,
                                    trabajos=[trabajo]
                                ))
                    
                    # Iniciar nuevo equipo
                    current_equipment = equipo_val
                    current_date = fecha_val if fecha_val else date.today()
                    current_repuestos = []
                    current_mano_obra = []
                    
                    # Agregar repuesto y mano de obra de esta fila si existen
                    if repuesto_val:
                        current_repuestos.append(repuesto_val)
                    if mano_obra_val:
                        current_mano_obra.append(mano_obra_val)
                
                # Si hay una nueva fecha (sin equipo), crear trabajo con datos acumulados
                elif fecha_val and current_equipment:
                    # Guardar trabajo anterior
                    if current_repuestos or current_mano_obra:
                        trabajo = self._create_work_entry(current_date, current_repuestos, current_mano_obra)
                        if trabajo:
                            # Buscar equipo existente
                            existing_eq = None
                            for eq in equipment_list:
                                if eq.nombre == current_equipment:
                                    existing_eq = eq
                                    break
                            
                            if existing_eq:
                                existing_eq.trabajos.append(trabajo)
                            else:
                                equipment_list.append(EquipmentData(
                                    propietario=sheet_name,
                                    nombre=current_equipment,
                                    trabajos=[trabajo]
                                ))
                    
                    # Iniciar nueva fecha
                    current_date = fecha_val
                    current_repuestos = []
                    current_mano_obra = []
                    
                    # Agregar repuesto y mano de obra de esta fila
                    if repuesto_val:
                        current_repuestos.append(repuesto_val)
                    if mano_obra_val:
                        current_mano_obra.append(mano_obra_val)
                
                # Solo repuestos/mano de obra (continuar acumulando)
                elif current_equipment and (repuesto_val or mano_obra_val):
                    if repuesto_val:
                        current_repuestos.append(repuesto_val)
                    if mano_obra_val:
                        current_mano_obra.append(mano_obra_val)
                
                # Fila completamente vacía - podría ser fin de equipo
                elif not any([equipo_val, fecha_val, repuesto_val, mano_obra_val]):
                    # Si tenemos datos acumulados, crear trabajo
                    if current_equipment and (current_repuestos or current_mano_obra):
                        trabajo = self._create_work_entry(current_date, current_repuestos, current_mano_obra)
                        if trabajo:
                            # Buscar equipo existente
                            existing_eq = None
                            for eq in equipment_list:
                                if eq.nombre == current_equipment:
                                    existing_eq = eq
                                    break
                            
                            if existing_eq:
                                existing_eq.trabajos.append(trabajo)
                            else:
                                equipment_list.append(EquipmentData(
                                    propietario=sheet_name,
                                    nombre=current_equipment,
                                    trabajos=[trabajo]
                                ))
                        
                        # Reset para próximo equipo
                        current_repuestos = []
                        current_mano_obra = []
            
            # Procesar último equipo si queda pendiente
            if current_equipment and (current_repuestos or current_mano_obra):
                trabajo = self._create_work_entry(current_date, current_repuestos, current_mano_obra)
                if trabajo:
                    existing_eq = None
                    for eq in equipment_list:
                        if eq.nombre == current_equipment:
                            existing_eq = eq
                            break
                    
                    if existing_eq:
                        existing_eq.trabajos.append(trabajo)
                    else:
                        equipment_list.append(EquipmentData(
                            propietario=sheet_name,
                            nombre=current_equipment,
                            trabajos=[trabajo]
                        ))
            
            logger.info(f"Procesados {len(equipment_list)} equipos en '{sheet_name}'")
            return equipment_list
            
        except Exception as e:
            logger.error(f"Error procesando hoja '{sheet_name}': {e}")
            self.errors.append(f"Error en hoja '{sheet_name}': {e}")
            return []
    
    def _find_header_row(self, worksheet) -> Optional[int]:
        """Encontrar la fila que contiene los headers"""
        max_row = min(worksheet.max_row, 10)  # Buscar en las primeras 10 filas
        
        for row_num in range(1, max_row + 1):
            row_values = []
            for col_num in range(1, 5):  # Primeras 4 columnas
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value:
                    row_values.append(str(cell_value).upper())
            
            # Verificar si esta fila contiene los headers esperados
            if len(row_values) >= 2:
                row_text = ' '.join(row_values)
                if 'EQUIPO' in row_text and ('FECHA' in row_text or 'REPUESTO' in row_text):
                    return row_num
        
        return None
    
    def _create_work_entry(self, fecha: date, repuestos: List[str], mano_obra: List[str]) -> Optional[WorkEntry]:
        """Crear una entrada de trabajo combinando repuestos y mano de obra"""
        if not repuestos and not mano_obra:
            return None
        
        # Combinar descripción
        descripcion_parts = []
        
        if repuestos:
            descripcion_parts.append("REPUESTOS: " + " | ".join(repuestos))
        
        if mano_obra:
            descripcion_parts.append("MANO DE OBRA: " + " | ".join(mano_obra))
        
        descripcion = " // ".join(descripcion_parts)
        
        # Extraer costo
        costo = self._extract_cost(descripcion)
        
        return WorkEntry(
            fecha=fecha if fecha else date.today(),
            descripcion=descripcion,
            presupuesto=costo
        )
    
    def _clean_string(self, text) -> str:
        """Limpiar y normalizar texto"""
        if text is None:
            return ""
        
        text = str(text).strip()
        if text.lower() in ['nan', 'none', '']:
            return ""
        
        return text
    
    def _parse_date(self, date_value) -> Optional[date]:
        """Parsear fecha con múltiples formatos"""
        if date_value is None:
            return None
        
        # Si ya es un objeto datetime
        if isinstance(date_value, (datetime, date)):
            return date_value.date() if isinstance(date_value, datetime) else date_value
        
        date_str = str(date_value).strip()
        if not date_str or date_str.lower() in ['nan', 'none', '']:
            return None
        
        # Formatos de fecha a probar
        date_formats = [
            '%Y-%m-%d %H:%M:%S',  # 2023-05-29 00:00:00
            '%Y-%m-%d',           # 2023-05-29
            '%d/%m/%Y',           # 29/05/2023
            '%d-%m-%Y',           # 29-05-2023
            '%d//%m//%Y',         # 14//6/2023
            '%d/%m/%y',           # 29/05/23
            '%d-%m-%y',           # 29-05-23
        ]
        
        # Intentar parsear con diferentes formatos
        for fmt in date_formats:
            try:
                parsed = datetime.strptime(date_str, fmt)
                return parsed.date()
            except ValueError:
                continue
        
        # Intentar con pandas si está disponible
        if PANDAS_AVAILABLE:
            try:
                return pd.to_datetime(date_str, dayfirst=True).date()
            except:
                pass
        
        logger.warning(f"No se pudo parsear fecha: {date_str}")
        return None
    
    def _extract_cost(self, text: str) -> float:
        """Extraer costo del texto"""
        # Buscar patrones de dinero
        money_patterns = [
            r'\$\s*(\d+(?:[.,]\d+)*)',  # $1000 o $1,000.50
            r'(\d+(?:[.,]\d+)*)\s*\$',  # 1000$ o 1,000.50$
            r'(\d+(?:[.,]\d+)*)\s*pesos',  # 1000 pesos
        ]
        
        for pattern in money_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                try:
                    # Tomar el primer match y convertir
                    amount_str = matches[0].replace(',', '').replace('.', '')
                    return float(amount_str)
                except ValueError:
                    continue
        
        return 0.0

# Función de compatibilidad
def validate_excel_file_final(file_path: str) -> Dict[str, Any]:
    """Validar archivo Excel con el parser final"""
    parser = ExcelParserFinal()
    equipment_data = parser.parse_excel_file(file_path)
    
    total_records = sum(len(eq.trabajos) for eq in equipment_data)
    
    return {
        'valid': len(equipment_data) > 0,
        'summary': {
            'total_equipment': len(equipment_data),
            'estimated_records': total_records,
            'sheets_processed': len(set(eq.propietario for eq in equipment_data))
        },
        'equipment_data': equipment_data,
        'errors': parser.errors,
        'warnings': parser.warnings
    }
