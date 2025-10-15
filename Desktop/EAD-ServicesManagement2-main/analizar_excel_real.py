"""
Análisis detallado del Excel real para entender su estructura
"""
import pandas as pd
import openpyxl
import numpy as np

def analizar_excel_detallado(archivo):
    print("="*80)
    print(f"ANÁLISIS DETALLADO DE: {archivo}")
    print("="*80)
    
    try:
        # Abrir con openpyxl para análisis más detallado
        workbook = openpyxl.load_workbook(archivo)
        
        for sheet_name in workbook.sheetnames:
            print(f"\n{'='*60}")
            print(f"HOJA: {sheet_name}")
            print(f"{'='*60}")
            
            worksheet = workbook[sheet_name]
            
            # Obtener dimensiones reales
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            print(f"Dimensiones: {max_row} filas x {max_col} columnas")
            
            # Analizar primeras 20 filas para entender estructura
            print(f"\nPRIMERAS 20 FILAS:")
            print("-" * 40)
            
            for row_num in range(1, min(21, max_row + 1)):
                row_data = []
                for col_num in range(1, min(max_col + 1, 10)):  # Primeras 10 columnas
                    cell = worksheet.cell(row=row_num, column=col_num)
                    value = cell.value
                    if value is not None:
                        value_str = str(value)[:30]  # Truncar a 30 chars
                        row_data.append(f"'{value_str}'")
                    else:
                        row_data.append("None")
                
                print(f"Fila {row_num:2d}: {' | '.join(row_data)}")
            
            # Buscar patrones de equipos
            print(f"\nBUSCANDO PATRONES DE EQUIPOS:")
            print("-" * 40)
            
            equipos_encontrados = []
            for row_num in range(1, min(max_row + 1, 100)):  # Primeras 100 filas
                for col_num in range(1, min(max_col + 1, 5)):  # Primeras 5 columnas
                    cell = worksheet.cell(row=row_num, column=col_num)
                    if cell.value:
                        value_str = str(cell.value).strip()
                        # Buscar patrones que parezcan equipos
                        if any(keyword in value_str.upper() for keyword in 
                              ['GRÚA', 'GRUA', 'HIDRO', 'MÓVIL', 'MOVIL', 'CAMIÓN', 'CAMION', 
                               'TRACTOR', 'EXCAVADORA', 'CATERPILLAR', 'SCANIA', 'VOLVO']):
                            equipos_encontrados.append({
                                'fila': row_num,
                                'columna': col_num,
                                'valor': value_str
                            })
            
            print(f"Equipos potenciales encontrados: {len(equipos_encontrados)}")
            for eq in equipos_encontrados[:10]:  # Mostrar primeros 10
                print(f"  Fila {eq['fila']}, Col {eq['columna']}: {eq['valor']}")
            
            # Buscar patrones de fechas
            print(f"\nBUSCANDO PATRONES DE FECHAS:")
            print("-" * 40)
            
            fechas_encontradas = []
            for row_num in range(1, min(max_row + 1, 100)):
                for col_num in range(1, min(max_col + 1, 10)):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    if cell.value:
                        value_str = str(cell.value)
                        # Buscar patrones de fecha
                        if any(pattern in value_str for pattern in 
                              ['/', '-', '2023', '2024', '2025']) and len(value_str) < 20:
                            fechas_encontradas.append({
                                'fila': row_num,
                                'columna': col_num,
                                'valor': value_str
                            })
            
            print(f"Fechas potenciales encontradas: {len(fechas_encontradas)}")
            for fecha in fechas_encontradas[:10]:
                print(f"  Fila {fecha['fila']}, Col {fecha['columna']}: {fecha['valor']}")
            
            # Analizar con pandas también
            print(f"\nANÁLISIS CON PANDAS:")
            print("-" * 40)
            
            try:
                df = pd.read_excel(archivo, sheet_name=sheet_name)
                print(f"Pandas shape: {df.shape}")
                print(f"Columnas: {list(df.columns)}")
                
                # Mostrar tipos de datos
                print(f"\nTipos de datos:")
                for col in df.columns:
                    dtype = df[col].dtype
                    non_null = df[col].count()
                    print(f"  {col}: {dtype} ({non_null} valores no nulos)")
                
                # Mostrar primeras filas no vacías
                df_clean = df.dropna(how='all')
                print(f"\nPrimeras 5 filas no vacías:")
                for i, (idx, row) in enumerate(df_clean.head().iterrows()):
                    print(f"  Fila {idx}: {dict(row)}")
                    if i >= 4:
                        break
                        
            except Exception as e:
                print(f"Error con pandas: {e}")
        
        workbook.close()
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    import os
    archivos = [f for f in os.listdir('.') if f.endswith('.xlsx')]
    
    for archivo in archivos:
        if 'Equipos' in archivo:  # Enfocarse en el archivo principal
            analizar_excel_detallado(archivo)
            break
